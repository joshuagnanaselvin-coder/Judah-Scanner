#!/usr/bin/env python3
"""
audit.py — Signal Quality Auditor for Judah Scanner.

Reads signal_log.xlsx, VERIFIES each trade by pulling Binance market data
and checking if TP was hit before SL (or vice versa), then analyzes all
verified outcomes and generates quality_report.html.

Run: python audit.py          (live verification via Binance API)
Run: python audit.py --offline (skip Binance, use only Excel data)

Verification logic:
  For each signal with Entry/SL/TP:
    1. Fetch klines from Binance starting at signal timestamp
    2. Walk candle-by-candle: did TP get hit before SL?
    3. Record: WON (TP hit first), LOST (SL hit first), EXPIRED (neither)
    4. Compute achieved RR, bars to target, actual PnL%

OFFLINE MODE:
  If Binance API is unreachable, audit.py falls back to analyzing
  whatever outcome data exists in the Trade Outcomes sheet.
"""

import os
import sys
import json
import math
import asyncio
import ssl
from datetime import datetime, timezone
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import aiohttp
except ImportError:
    print("ERROR: aiohttp not installed. Run: pip install aiohttp")
    sys.exit(1)

# ============================================================
# CONSTANTS
# ============================================================

BINANCE_REST = "https://api.binance.com/api/v3/klines"
FETCH_TIMEOUT = 15
VERIFY_WINDOW_HOURS = {"1h": 168, "4h": 504, "1d": 504, "1m": 24, "5m": 48, "15m": 72, "30m": 96}

# ============================================================
# HELPERS
# ============================================================

def _safe_get(row, col_idx):
    try:
        v = row[col_idx]
        return None if (v is None or (isinstance(v, str) and v.strip() == "")) else v
    except (IndexError, TypeError):
        return None


def _safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            cleaned = str(v).strip().replace("$", "").replace(",", "").replace("%", "").replace(" ", "")
            return float(cleaned)
        except (TypeError, ValueError):
            return None


def _safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_timestamp(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        if v > 1e12:
            return datetime.fromtimestamp(v / 1000, tz=timezone.utc)
        return datetime.fromtimestamp(v, tz=timezone.utc)
    s = str(v).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                "%m/%d/%Y %H:%M", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        n = float(s)
        if n > 1e12:
            return datetime.fromtimestamp(n / 1000, tz=timezone.utc)
        return datetime.fromtimestamp(n, tz=timezone.utc)
    except (ValueError, TypeError, OSError):
        return None


def _pct(num, den):
    if not den:
        return 0.0
    return round(num / den * 100, 1)


def _avg(vals):
    nums = [v for v in vals if v is not None]
    if not nums:
        return 0.0
    return round(sum(nums) / len(nums), 2)


def _find_col(headers, candidates):
    headers_lower = [h.lower().strip() if h else "" for h in headers]
    for cand in candidates:
        cand_lower = cand.lower().strip()
        for i, h in enumerate(headers_lower):
            if cand_lower in h or h in cand_lower:
                return i
    return None


def _auto_detect_cols(headers, schema):
    mapping = {}
    for key, candidates in schema.items():
        idx = _find_col(headers, candidates)
        if idx is not None:
            mapping[key] = idx
    return mapping


# ============================================================
# BINANCE API FETCHER (aiohttp — async batch)
# ============================================================

def _normalize_symbol(symbol: str) -> str:
    """Convert any format to Binance: ETH/USDT → ETHUSDT, btc → BTCUSDT"""
    s = symbol.upper().strip()
    s = s.replace("/USDT", "").replace("-USDT", "").replace("_USDT", "")
    if s.endswith("USD"):
        s = s[:-3]
    if not s.endswith("USDT"):
        s = s + "USDT"
    return s


def _tf_to_interval(timeframe: str) -> str:
    """Convert 1H/4h/1D → 1h/4h/1d"""
    tf = timeframe.strip().upper()
    mapping = {"1H": "1h", "4H": "4h", "1D": "1d",
               "1M": "1m", "5M": "5m", "15M": "15m", "30M": "30m"}
    return mapping.get(tf, tf.lower())


def _kline_dict(k):
    return {"time": k[0], "open": float(k[1]), "high": float(k[2]),
            "low": float(k[3]), "close": float(k[4]), "volume": float(k[5]),
            "close_time": k[6]}


async def _fetch_klines(session, symbol, interval, start_ms, end_ms):
    """Single kline fetch via aiohttp."""
    sym = _normalize_symbol(symbol)
    url = (f"{BINANCE_REST}?symbol={sym}&interval={interval}"
           f"&startTime={int(start_ms)}&endTime={int(end_ms)}&limit=1000")
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=FETCH_TIMEOUT)) as resp:
            if resp.status != 200:
                return None, f"HTTP {resp.status}"
            data = await resp.json()
            if not isinstance(data, list):
                return None, str(data.get("msg", "bad response"))
            if not data:
                return [], None
            return [_kline_dict(k) for k in data], None
    except Exception as e:
        return None, str(e)[:120]


async def _batch_fetch_klines(jobs):
    """Fetch klines for all jobs in parallel. Single aiohttp session.

    Args:
        jobs: list of (index, symbol, direction, entry, sl, tp, ts, tf, tf_interval)
    Returns:
        dict: index → list of kline dicts (empty list if failed)
    """
    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE
    connector = aiohttp.TCPConnector(ssl=ssl_ctx, limit=10)

    async with aiohttp.ClientSession(
        connector=connector,
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    ) as session:
        # Connectivity check
        try:
            async with session.get(
                f"{BINANCE_REST.replace('/klines', '')}/time",
                timeout=aiohttp.ClientTimeout(total=8)
            ) as resp:
                if resp.status != 200:
                    print(f"  [WARN] Binance API HTTP {resp.status} — check connectivity")
                    return {i: [] for i, *_ in jobs}
        except Exception as e:
            print(f"  [WARN] Cannot reach Binance API: {str(e)[:100]}")
            print("  TIP: Run with --offline flag:  python audit.py --offline")
            return {i: [] for i, *_ in jobs}

        # Build fetch tasks with semaphore (max 5 concurrent)
        semaphore = asyncio.Semaphore(5)
        results = {}

        async def fetch_one(idx, sym, direction, entry, sl, tp, ts, tf, tf_int):
            tf_interval = _tf_to_interval(tf_int)
            window_hours = VERIFY_WINDOW_HOURS.get(tf_interval, 168)
            start_ms = int(ts * 1000) if ts < 1e12 else int(ts)
            end_ms = start_ms + window_hours * 3600 * 1000
            start_ms = start_ms - 60000  # 1 min buffer before signal

            async with semaphore:
                klines, err = await _fetch_klines(session, sym, tf_interval, start_ms, end_ms)
            if err:
                results[idx] = None  # mark as failed
                print(f"  [fetch FAIL] {sym} {tf_interval}: {err}")
            else:
                results[idx] = klines if klines is not None else []

        tasks = [fetch_one(*job) for job in jobs]
        for coro in asyncio.as_completed(tasks):
            await coro

    return results


# ============================================================
# TRADE WALKER — candle-by-candle TP/SL check
# ============================================================

def _verify_from_klines(direction, entry, sl, tp, klines):
    """Walk klines candle-by-candle to find which target was hit first."""
    if not klines or len(klines) < 2:
        return None

    is_bullish = direction.upper() == "BULLISH"
    best_for = entry
    worst_against = entry
    bars_elapsed = 0

    for k in klines:
        if bars_elapsed == 0:
            bars_elapsed = 1
            continue

        high = k["high"]
        low = k["low"]

        if is_bullish:
            # Bullish: price goes up → TP first = WIN, SL first = LOSS
            if high >= tp:
                return {
                    "result": "WON",
                    "exit_price": tp,
                    "exit_bar": bars_elapsed,
                    "achieved_rr": round((tp - entry) / (entry - sl), 2) if entry != sl else 0,
                    "pnl_pct": round((tp - entry) / entry * 100, 2),
                    "exit_reason": f"TP hit bar {bars_elapsed}",
                    "high_water": round(max(best_for, high), 2),
                    "low_water": round(min(worst_against, low), 2),
                }
            if low <= sl:
                return {
                    "result": "LOST",
                    "exit_price": sl,
                    "exit_bar": bars_elapsed,
                    "achieved_rr": 0,
                    "pnl_pct": round((sl - entry) / entry * 100, 2),
                    "exit_reason": f"SL hit bar {bars_elapsed}",
                    "high_water": round(max(best_for, high), 2),
                    "low_water": round(min(worst_against, low), 2),
                }
            if high > best_for:
                best_for = high
            if low < worst_against:
                worst_against = low
        else:
            # Bearish: price goes down → SL first = LOSS, TP first = WIN
            if low <= sl:
                return {
                    "result": "LOST",
                    "exit_price": sl,
                    "exit_bar": bars_elapsed,
                    "achieved_rr": 0,
                    "pnl_pct": round((entry - sl) / entry * 100, 2),
                    "exit_reason": f"SL hit bar {bars_elapsed}",
                    "high_water": round(max(best_for, high), 2),
                    "low_water": round(min(worst_against, low), 2),
                }
            if high >= tp:
                return {
                    "result": "WON",
                    "exit_price": tp,
                    "exit_bar": bars_elapsed,
                    "achieved_rr": round((entry - tp) / (sl - entry), 2) if entry != sl else 0,
                    "pnl_pct": round((entry - tp) / entry * 100, 2),
                    "exit_reason": f"TP hit bar {bars_elapsed}",
                    "high_water": round(max(best_for, high), 2),
                    "low_water": round(min(worst_against, low), 2),
                }
            if high > best_for:
                best_for = high
            if low < worst_against:
                worst_against = low

        bars_elapsed += 1

    # Neither target hit
    last_close = klines[-1]["close"]
    return {
        "result": "EXPIRED",
        "exit_price": last_close,
        "exit_bar": bars_elapsed,
        "achieved_rr": 0,
        "pnl_pct": round((last_close - entry) / entry * 100, 2),
        "exit_reason": f"No target in {bars_elapsed} bars",
        "high_water": round(best_for, 2),
        "low_water": round(worst_against, 2),
    }


# ============================================================
# DATA LOADING
# ============================================================

def load_workbook(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: {filepath} not found in {os.getcwd()}")
        sys.exit(1)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    return wb


def read_sheet(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


# ============================================================
# TRADE OBJECT
# ============================================================

class Trade:
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)

    def get(self, key, default=None):
        return getattr(self, key, default)

    def to_dict(self):
        return self.__dict__


# ============================================================
# ANALYSIS CONTAINER
# ============================================================

class AuditResult:
    def __init__(self):
        self.total_signals = 0
        self.verified = 0
        self.skipped = 0
        self.fetch_failed = 0
        self.no_data = 0
        self.from_excel = 0  # outcomes from Trade Outcomes sheet

        self.wins = 0
        self.losses = 0
        self.expired = 0
        self.gross_profit = 0.0
        self.gross_loss = 0.0
        self.all_pnl = []
        self.all_rr = []
        self.all_bars = []
        self.max_consec_wins = 0
        self.max_consec_losses = 0
        self.best_trade = None
        self.worst_trade = None

        self.by_tier = defaultdict(list)
        self.by_session = defaultdict(list)
        self.by_tf = defaultdict(list)
        self.by_direction = defaultdict(list)
        self.by_score_range = defaultdict(list)
        self.by_confluence = defaultdict(list)
        self.by_freshness = defaultdict(list)
        self.by_bars = defaultdict(list)
        self.by_coin = defaultdict(list)

        # CRT
        self.crt_displacement = {"high": [], "low": []}
        self.crt_ote = {"in_ote": [], "not_ote": []}
        self.crt_optimal = {"optimal": [], "not_optimal": []}
        self.crt_range_break = {"break": [], "no_break": []}
        self.crt_zone = defaultdict(list)
        self.crt_retracement = {"618_70": [], "50_618": [], "other": []}

        # SMC
        self.smc_ob = {"has_ob": [], "no_ob": []}
        self.smc_ob_touches = defaultdict(list)
        self.smc_fvg = {"has_fvg": [], "no_fvg": []}
        self.smc_msb = {"has_msb": [], "no_msb": []}
        self.smc_liquidity = {"swept": [], "not_swept": []}
        self.boosts = {"has_boost": [], "no_boost": []}

        self.combos = defaultdict(list)
        self.daily = defaultdict(list)
        self.equity_curve = []


# ============================================================
# RECORD TRADE INTO ALL BUCKETS
# ============================================================

def _record(trade, ar):
    result_str = trade.result
    pnl_f = trade.pnl_pct
    rr_f = trade.achieved_rr

    ar.verified += 1
    ar.all_pnl.append(pnl_f)
    ar.all_rr.append(rr_f)
    if trade.exit_bar:
        ar.all_bars.append(trade.exit_bar)

    if result_str == "WON":
        ar.wins += 1
        ar.gross_profit += abs(pnl_f) if pnl_f else 0
        if ar.best_trade is None or (pnl_f or -999) > (ar.best_trade.pnl_pct or -999):
            ar.best_trade = trade
    elif result_str == "LOST":
        ar.losses += 1
        ar.gross_loss += abs(pnl_f) if pnl_f else 0
        if ar.worst_trade is None or (pnl_f or 999) < (ar.worst_trade.pnl_pct or 999):
            ar.worst_trade = trade
    else:
        ar.expired += 1

    # Buckets
    result_dict = {"result": result_str, "pnl_pct": pnl_f, "won": result_str == "WON",
                   "lost": result_str == "LOST", "breakeven": result_str == "EXPIRED"}

    tier = (trade.tier or "UNKNOWN").upper()
    ar.by_tier[tier].append(result_dict)

    session = (trade.session or "UNKNOWN").upper()
    ar.by_session[session].append(result_dict)

    tf = (trade.timeframe or trade.engine or "UNKNOWN").upper()
    ar.by_tf[tf].append(result_dict)

    direction = (trade.direction or "UNKNOWN").upper()
    ar.by_direction[direction].append(result_dict)

    symbol = (trade.symbol or "UNKNOWN").upper()
    ar.by_coin[symbol].append(result_dict)

    score = trade.composite_score
    if score is None:
        score = getattr(trade, "base_score", None)
    if score is not None:
        s = float(score)
        if s >= 80: bucket = "80-100"
        elif s >= 70: bucket = "70-79"
        elif s >= 60: bucket = "60-69"
        elif s >= 50: bucket = "50-59"
        elif s >= 40: bucket = "40-49"
        elif s >= 30: bucket = "30-39"
        else: bucket = "<30"
        ar.by_score_range[bucket].append(result_dict)

    cc = trade.confluence_count if trade.confluence_count is not None else 0
    ck = str(cc) if cc < 3 else "3+"
    ar.by_confluence[ck].append(result_dict)

    fr = (trade.freshness_at_entry or "UNKNOWN").lower()
    ar.by_freshness[fr].append(result_dict)

    if trade.exit_bar is not None:
        if trade.exit_bar <= 2:
            ar.by_bars["1-2 bars"].append(result_dict)
        elif trade.exit_bar <= 8:
            ar.by_bars["3-8 bars"].append(result_dict)
        elif trade.exit_bar <= 24:
            ar.by_bars["9-24 bars"].append(result_dict)
        else:
            ar.by_bars["25+ bars"].append(result_dict)

    # CRT
    dr = getattr(trade, "displacement_ratio", None)
    if dr is not None:
        key = "high" if float(dr) >= 2.0 else "low"
        ar.crt_displacement[key].append(result_dict)

    iote = getattr(trade, "in_ote", None)
    if iote is not None:
        val = str(iote).lower().strip()
        if val in ("yes", "true", "1"):
            ar.crt_ote["in_ote"].append(result_dict)
            opt = str(getattr(trade, "in_optimal_ote", "") or "").lower().strip()
            ar.crt_optimal["optimal" if opt in ("yes", "true", "1") else "not_optimal"].append(result_dict)
        else:
            ar.crt_ote["not_ote"].append(result_dict)
            ar.crt_optimal["not_optimal"].append(result_dict)

    rb = getattr(trade, "range_break", None)
    if rb is not None:
        val = str(rb).lower().strip()
        ar.crt_range_break["break" if val in ("yes", "true", "1") else "no_break"].append(result_dict)

    zone = (getattr(trade, "premium_discount", "UNKNOWN") or "UNKNOWN").upper()
    ar.crt_zone[zone].append(result_dict)

    retr = getattr(trade, "retracement_pct", None)
    if retr is not None:
        r = float(retr)
        if 61.8 <= r <= 70:
            ar.crt_retracement["618_70"].append(result_dict)
        elif 50 <= r < 61.8:
            ar.crt_retracement["50_618"].append(result_dict)
        else:
            ar.crt_retracement["other"].append(result_dict)

    # SMC
    ob_type = getattr(trade, "ob_type", None)
    if ob_type and str(ob_type).lower() not in ("none", "n/a", ""):
        ar.smc_ob["has_ob"].append(result_dict)
        touches = getattr(trade, "ob_touches", 1) or 1
        ar.smc_ob_touches[str(touches)].append(result_dict)
    else:
        ar.smc_ob["no_ob"].append(result_dict)

    fvg_type = getattr(trade, "fvg_type", None)
    if fvg_type and str(fvg_type).lower() not in ("none", "n/a", ""):
        ar.smc_fvg["has_fvg"].append(result_dict)
    else:
        ar.smc_fvg["no_fvg"].append(result_dict)

    msb = getattr(trade, "msb_confirmed", None)
    if msb is not None:
        val = str(msb).lower().strip()
        ar.smc_msb["has_msb" if val in ("yes", "true", "1") else "no_msb"].append(result_dict)

    liq = getattr(trade, "liquidity_swept", None)
    if liq is not None:
        val = str(liq).lower().strip()
        ar.smc_liquidity["swept" if val in ("yes", "true", "1") else "not_swept"].append(result_dict)

    boosts = getattr(trade, "priority_boosts", "") or ""
    if str(boosts).strip() and str(boosts).strip() not in ("N/A", "None", ""):
        ar.boosts["has_boost"].append(result_dict)
    else:
        ar.boosts["no_boost"].append(result_dict)

    # Combos
    combos = [
        (tier, session), (tier, tf), (direction, session),
        (f"score_{bucket}", tf), (f"confluence_{ck}", fr),
        (symbol, tf), (symbol, tier),
    ]
    for c in combos:
        if c:
            ar.combos[c].append(result_dict)

    # Daily
    ts = _parse_timestamp(getattr(trade, "timestamp", None))
    if ts:
        day = ts.strftime("%Y-%m-%d")
        ar.daily[day].append({"won": result_str == "WON", "lost": result_str == "LOST",
                               "pnl": pnl_f or 0.0})


# ============================================================
# STATS HELPERS
# ============================================================

def _bucket_stats(bucket_list):
    n = len(bucket_list)
    if n == 0:
        return {"count": 0, "wr": 0, "avg_rr": 0, "avg_pnl": 0}
    won = sum(1 for t in bucket_list if t.get("won"))
    lost = sum(1 for t in bucket_list if t.get("lost"))
    wr = _pct(won, won + lost) if (won + lost) > 0 else 0.0
    return {"count": n, "won": won, "lost": lost, "wr": wr, "avg_rr": 0, "avg_pnl": 0}


# ============================================================
# SIGNAL LOADING
# ============================================================

def load_signals_from_excel(wb):
    """Load signals + existing trade outcomes from Excel."""
    ws_signals = None
    ws_trade_outcomes = None

    for name in wb.sheetnames:
        nl = name.lower().strip()
        if nl in ("signals", "signal"):
            ws_signals = wb[name]
        elif "trade" in nl:
            ws_trade_outcomes = wb[name]

    if not ws_signals:
        print("ERROR: No 'Signals' sheet found.")
        sys.exit(1)

    sig_rows = read_sheet(ws_signals)
    if not sig_rows:
        print("ERROR: Signals sheet is empty.")
        sys.exit(1)

    headers = sig_rows[0]
    print(f"Signals headers ({len(headers)}): {headers}")

    schema = {
        "signal_id": ["Signal ID", "signal id", "id"],
        "timestamp": ["Timestamp", "timestamp", "time"],
        "symbol": ["Symbol", "symbol", "pair"],
        "timeframe": ["Timeframe", "timeframe", "engine", "tf"],
        "direction": ["Direction", "direction"],
        "tier": ["Tier", "tier"],
        "entry": ["Entry Price", "entry price", "entry"],
        "stop_loss": ["Stop Loss", "stop loss", "sl"],
        "take_profit": ["Take Profit", "take profit", "tp"],
        "composite_score": ["Composite Score", "composite score"],
        "base_score": ["Base Score", "base score", "score at birth"],
        "session": ["Session", "session"],
        "confluence_count": ["Confluence Count", "confluence count"],
        "priority_boosts": ["Priority Boosts", "priority boosts"],
        "freshness_state": ["Freshness State", "freshness state", "freshness at entry"],
        "status": ["Status", "status"],
        "displacement_ratio": ["Displacement Ratio", "displacement ratio"],
        "in_ote": ["In OTE", "in ote"],
        "in_optimal_ote": ["In Optimal OTE", "in optimal ote"],
        "range_break": ["Range Break", "range break"],
        "premium_discount": ["Premium/Discount", "premium/discount", "zone"],
        "retracement_pct": ["Retracement %", "retracement %"],
        "ob_type": ["OB Type", "ob type"],
        "ob_touches": ["OB Touches", "OB Retest Count", "ob touches"],
        "fvg_type": ["FVG Type", "fvg type"],
        "liquidity_swept": ["Liquidity Swept", "liquidity swept"],
        "msb_confirmed": ["MSB Confirmed", "msb confirmed", "msb type"],
    }
    cols = _auto_detect_cols(headers, schema)

    # Load trade outcomes if Trade Levels / Trade Outcomes sheet exists
    trade_outcomes = {}
    if ws_trade_outcomes:
        to_rows = read_sheet(ws_trade_outcomes)
        if to_rows:
            to_headers = to_rows[0]
            print(f"Trade sheet headers ({len(to_headers)}): {to_headers}")
            to_schema = {
                "result": ["Result", "result", "Outcome"],
                "signal_id": ["Signal ID", "signal id", "id"],
                "pnl_pct": ["PnL %", "pnl %", "pnl_pct", "closed pnl %"],
                "pnl_dollar": ["PnL $", "pnl $"],
                "achieved_rr": ["Achieved RR", "achieved rr"],
                "exit_reason": ["Exit Reason", "exit reason"],
                "exit_price": ["Exit Price", "exit price", "closed at price"],
                "close_timestamp": ["Close Timestamp", "close timestamp", "close time"],
            }
            to_cols = _auto_detect_cols(to_headers, to_schema)
            for r in to_rows[1:]:
                sid = _safe_get(r, to_cols.get("signal_id")) if "signal_id" in to_cols else None
                if sid:
                    trade_outcomes[sid] = {k: _safe_get(r, ci) for k, ci in to_cols.items() if ci is not None}
            print(f"  Loaded {len(trade_outcomes)} trade outcomes from Trade Levels sheet")

    signals = []
    for r in sig_rows[1:]:
        sig = {}
        for key, ci in cols.items():
            v = _safe_get(r, ci) if ci is not None else None
            if v is not None:
                sig[key] = v
        sid = sig.get("signal_id")
        if sid and sid in trade_outcomes:
            sig["_trade_outcome"] = trade_outcomes[sid]
        signals.append(sig)

    return signals


# ============================================================
# VERIFY SIGNALS
# ============================================================

def verify_signals(signals, offline=False):
    """Verify each signal's outcome using Binance data.

    Priority: Trade Outcomes from Excel > Binance live verification > skip
    """
    print(f"\nVerifying {len(signals)} signals...")
    if offline:
        print("  OFFLINE MODE — using Trade Outcomes from Excel only")

    verified = []
    skipped = 0
    fetch_failed = 0
    no_data = 0

    # Pass 1: Use cached outcomes from Excel, collect jobs needing Binance
    cached = []
    binance_jobs = []

    for i, sig in enumerate(signals):
        symbol = sig.get("symbol", "")
        if not symbol:
            skipped += 1
            continue

        # Always prefer cached outcome from Excel
        if "_trade_outcome" in sig:
            to = sig["_trade_outcome"]
            result_val = str(to.get("result", "")).upper()
            if result_val in ("WON", "LOST", "BREAKEVEN"):
                trade = _build_trade_from_signal(sig, to)
                cached.append(trade)
                continue

        # Need entry/SL/TP
        entry = _safe_float(sig.get("entry"))
        sl = _safe_float(sig.get("stop_loss"))
        tp = _safe_float(sig.get("take_profit"))

        if not entry or not sl or not tp:
            no_data += 1
            continue

        if offline:
            no_data += 1
            continue

        # Parse timestamp
        ts_val = sig.get("timestamp")
        signal_ts = _parse_timestamp(ts_val)
        if not signal_ts:
            signal_ts = datetime.now(timezone.utc)

        tf = (sig.get("timeframe") or sig.get("engine") or "4h").lower()
        binance_jobs.append((i, symbol, sig.get("direction", ""),
                              entry, sl, tp, signal_ts.timestamp(), tf, tf))

    # Pass 2: Batch fetch from Binance (single event loop)
    kline_map = {}
    if binance_jobs and not offline:
        print(f"  Fetching klines for {len(binance_jobs)} signals from Binance...")
        kline_map = asyncio.run(_batch_fetch_klines(binance_jobs))

    # Pass 3: Build Trade objects from kline verification
    for (idx, sym, direction, entry, sl, tp, signal_ts, tf, tf_interval) in binance_jobs:
        klines = kline_map.get(idx)
        if klines is None:
            fetch_failed += 1
            continue
        if not klines:
            fetch_failed += 1
            continue

        result = _verify_from_klines(direction, entry, sl, tp, klines)
        if result is None:
            fetch_failed += 1
            continue

        sig = signals[idx]
        trade = _build_trade_from_verification(sig, result, direction, entry, sl, tp, signal_ts, tf)
        verified.append(trade)

    verified.extend(cached)

    print(f"\n  Results: {len(verified)} verified ({len(cached)} from Excel, {len(verified)-len(cached)} from Binance)")
    print(f"  Skipped: {skipped} | Fetch failed: {fetch_failed} | No TP/SL data: {no_data}")

    verify_stats = {
        "skipped": skipped,
        "fetch_failed": fetch_failed,
        "no_data": no_data,
        "verified": len(verified),
        "from_excel": len(cached),
    }

    return verified, verify_stats


def _build_trade_from_signal(sig, to):
    """Build Trade from signal + existing Excel outcome."""
    symbol = sig.get("symbol", "")
    return Trade(
        symbol=symbol,
        timeframe=sig.get("timeframe", ""),
        engine=sig.get("engine", ""),
        direction=sig.get("direction", ""),
        tier=sig.get("tier", ""),
        result=str(to.get("result", "")).upper().strip(),
        entry=_safe_float(sig.get("entry")),
        stop_loss=_safe_float(sig.get("stop_loss")),
        take_profit=_safe_float(sig.get("take_profit")),
        pnl_pct=_safe_float(to.get("pnl_pct")),
        achieved_rr=_safe_float(to.get("achieved_rr")),
        exit_reason=to.get("exit_reason", ""),
        exit_price=_safe_float(to.get("exit_price")),
        exit_bar=None,
        timestamp=sig.get("timestamp"),
        close_timestamp=to.get("close_timestamp"),
        composite_score=_safe_float(sig.get("composite_score")) or _safe_float(sig.get("compositeScore")),
        base_score=sig.get("base_score"),
        crt_score=sig.get("crt_score"),
        smc_score=sig.get("smc_score"),
        session=sig.get("session"),
        confluence_count=_safe_int(sig.get("confluence_count")),
        priority_boosts=sig.get("priority_boosts", ""),
        freshness_at_entry=sig.get("freshness_state"),
        displacement_ratio=_safe_float(sig.get("displacement_ratio")),
        in_ote=sig.get("in_ote"),
        in_optimal_ote=sig.get("in_optimal_ote"),
        range_break=sig.get("range_break"),
        premium_discount=sig.get("premium_discount"),
        retracement_pct=_safe_float(sig.get("retracement_pct")),
        ob_type=sig.get("ob_type"),
        ob_touches=_safe_int(sig.get("ob_touches")),
        fvg_type=sig.get("fvg_type"),
        liquidity_swept=sig.get("liquidity_swept"),
        msb_confirmed=sig.get("msb_confirmed"),
    )


def _build_trade_from_verification(sig, result, direction, entry, sl, tp, signal_ts, tf):
    """Build Trade from Binance-verified outcome."""
    return Trade(
        symbol=sig.get("symbol", ""),
        timeframe=tf,
        engine=tf,
        direction=direction,
        tier=sig.get("tier", ""),
        result=result["result"],
        entry=entry,
        stop_loss=sl,
        take_profit=tp,
        pnl_pct=result["pnl_pct"],
        achieved_rr=result["achieved_rr"],
        exit_reason=result["exit_reason"],
        exit_price=result["exit_price"],
        exit_bar=result["exit_bar"],
        timestamp=signal_ts.timestamp(),
        composite_score=_safe_float(sig.get("composite_score")) or _safe_float(sig.get("compositeScore")),
        base_score=sig.get("base_score"),
        crt_score=sig.get("crt_score"),
        smc_score=sig.get("smc_score"),
        session=sig.get("session"),
        confluence_count=_safe_int(sig.get("confluence_count")),
        priority_boosts=sig.get("priority_boosts", ""),
        freshness_at_entry=sig.get("freshness_state"),
        displacement_ratio=_safe_float(sig.get("displacement_ratio")),
        in_ote=sig.get("in_ote"),
        in_optimal_ote=sig.get("in_optimal_ote"),
        range_break=sig.get("range_break"),
        premium_discount=sig.get("premium_discount"),
        retracement_pct=_safe_float(sig.get("retracement_pct")),
        ob_type=sig.get("ob_type"),
        ob_touches=_safe_int(sig.get("ob_touches")),
        fvg_type=sig.get("fvg_type"),
        liquidity_swept=sig.get("liquidity_swept"),
        msb_confirmed=sig.get("msb_confirmed"),
    )


# ============================================================
# ANALYSIS
# ============================================================

def analyze(trades):
    ar = AuditResult()
    ar.total_signals = len(trades)

    max_w, max_l, cur_w, cur_l = 0, 0, 0, 0

    for trade in trades:
        r = trade.result
        if r == "WON":
            cur_w += 1; cur_l = 0
            max_w = max(max_w, cur_w)
        elif r == "LOST":
            cur_l += 1; cur_w = 0
            max_l = max(max_l, cur_l)

    ar.max_consec_wins = max_w
    ar.max_consec_losses = max_l

    for trade in trades:
        _record(trade, ar)

    # Equity curve
    sorted_trades = sorted(trades, key=lambda t: t.timestamp or 0)
    cumulative = 0.0
    for t in sorted_trades:
        cumulative += (t.pnl_pct or 0.0)
        ts = _parse_timestamp(t.timestamp)
        if ts:
            ar.equity_curve.append({"date": ts.strftime("%Y-%m-%d"), "cumulative": round(cumulative, 2)})
            day = ts.strftime("%Y-%m-%d")
            ar.daily[day].append({"won": t.result == "WON", "lost": t.result == "LOST",
                                   "pnl": t.pnl_pct or 0.0})

    return ar


# ============================================================
# HTML GENERATION
# ============================================================

def generate_html(ar, all_trades, verify_stats=None):
    total_closed = ar.wins + ar.losses
    win_rate = _pct(ar.wins, total_closed) if total_closed > 0 else 0.0
    avg_rr = _avg(ar.all_rr)
    avg_pnl = _avg(ar.all_pnl)
    avg_bars = _avg(ar.all_bars)
    pf = round(ar.gross_profit / ar.gross_loss, 2) if ar.gross_loss > 0 else (999.0 if ar.gross_profit > 0 else 0.0)

    insufficient = total_closed < 20
    wr_text = "Collecting data... need 20+ verified trades for reliable analysis" if insufficient else ""

    all_ts = [t.timestamp for t in all_trades if t.timestamp]
    if all_ts:
        dmin = datetime.fromtimestamp(min(all_ts), tz=timezone.utc).strftime("%Y-%m-%d")
        dmax = datetime.fromtimestamp(max(all_ts), tz=timezone.utc).strftime("%Y-%m-%d")
        date_range = f"{dmin} to {dmax}"
    else:
        date_range = "N/A"

    # Source label
    from_excel = verify_stats.get("from_excel", 0) if verify_stats else 0
    from_binance = len(all_trades) - from_excel
    if from_excel > 0 and from_binance > 0:
        source_label = f"{from_binance} live-verified, {from_excel} from Excel"
        source_badge = 'method-badge method-cached'
    elif from_excel > 0:
        source_label = f"{from_excel} from Excel (Trade Outcomes)"
        source_badge = 'method-badge method-cached'
    else:
        source_label = f"{from_binance} live-verified via Binance API"
        source_badge = 'method-badge method-live'

    # ---- Breakdown stats ----
    def bstats(d):
        r = {}
        for k, v in d.items():
            n = len(v)
            if n == 0:
                continue
            won = sum(1 for t in v if t.get("won"))
            lost = sum(1 for t in v if t.get("lost"))
            wr_val = _pct(won, won + lost) if (won + lost) > 0 else 0.0
            r[k] = {"count": n, "won": won, "lost": lost, "wr": wr_val, "avg_pnl": 0, "avg_rr": 0}
        return r

    def breakdown_min3(d):
        return {k: v for k, v in d.items() if v["count"] >= 3}

    tier_stats = breakdown_min3(bstats(ar.by_tier))
    session_stats = breakdown_min3(bstats(ar.by_session))
    tf_stats = breakdown_min3(bstats(ar.by_tf))
    direction_stats = breakdown_min3(bstats(ar.by_direction))
    coin_stats = breakdown_min3(bstats(ar.by_coin))
    score_stats = breakdown_min3(bstats(ar.by_score_range))
    confluence_stats = breakdown_min3(bstats(ar.by_confluence))
    freshness_stats = breakdown_min3(bstats(ar.by_freshness))
    bars_stats = breakdown_min3(bstats(ar.by_bars))

    # CRT
    disp_h = bstats(ar.crt_displacement["high"])
    disp_l = bstats(ar.crt_displacement["low"])
    ote_i = bstats(ar.crt_ote["in_ote"])
    opt_i = bstats(ar.crt_optimal["optimal"])
    rb_y = bstats(ar.crt_range_break["break"])
    zone_stats = bstats(ar.crt_zone)
    retr_stats = bstats(ar.crt_retracement)

    # SMC
    ob_y = bstats(ar.smc_ob["has_ob"])
    ob_n = bstats(ar.smc_ob["no_ob"])
    fvg_y = bstats(ar.smc_fvg["has_fvg"])
    msb_y = bstats(ar.smc_msb["has_msb"])
    liq_y = bstats(ar.smc_liquidity["swept"])
    boost_y = bstats(ar.boosts["has_boost"])
    ob_touch_stats = bstats(ar.smc_ob_touches)

    # Combo stats
    combo_list = []
    for ck, trades_list in ar.combos.items():
        n = len(trades_list)
        if n < 3:
            continue
        won = sum(1 for t in trades_list if t.get("won"))
        lost = sum(1 for t in trades_list if t.get("lost"))
        wr_val = _pct(won, won + lost) if (won + lost) > 0 else 0.0
        combo_list.append({"combo": " + ".join(str(k) for k in ck), "wr": wr_val,
                           "count": n, "won": won, "lost": lost})
    combo_list.sort(key=lambda x: x["wr"], reverse=True)
    top5 = combo_list[:5]
    bottom3 = combo_list[-3:] if len(combo_list) >= 3 else []

    # Best/worst
    best = ar.best_trade
    worst = ar.worst_trade
    best_str = f"{best.symbol} {best.direction} {best.pnl_pct:+.1f}%" if best else "N/A"
    worst_str = f"{worst.symbol} {worst.direction} {worst.pnl_pct:+.1f}%" if worst else "N/A"

    # ---- Recommendations ----
    recommendations = []
    if insufficient:
        recommendations.append(f"<em>{wr_text}</em>")
    else:
        if session_stats:
            bs = max(session_stats.items(), key=lambda x: x[1]["wr"])
            ws_s = min(session_stats.items(), key=lambda x: x[1]["wr"])
            recommendations.append(f"Best session: <strong>{bs[0]}</strong> ({bs[1]['wr']}% WR vs {win_rate:.0f}% overall)")
            if ws_s[1]["wr"] < win_rate - 10:
                recommendations.append(f"Avoid <strong>{ws_s[0]}</strong> — only {ws_s[1]['wr']}% WR")

        if tf_stats:
            btf = max(tf_stats.items(), key=lambda x: x[1]["wr"])
            recommendations.append(f"Best timeframe: <strong>{btf[0]}</strong> ({btf[1]['wr']}% WR)")

        if score_stats:
            bsc = max(score_stats.items(), key=lambda x: x[1]["wr"])
            recommendations.append(f"Score sweet spot: <strong>{bsc[0]}</strong> ({bsc[1]['wr']}% WR)")

        if top5:
            c = top5[0]
            recommendations.append(f"Best combo: <strong>{c['combo']}</strong> → {c['wr']}% WR ({c['count']} trades)")
        if bottom3:
            c = bottom3[0]
            recommendations.append(f"Worst combo: <strong>{c['combo']}</strong> → {c['wr']}% WR — <strong>avoid</strong>")

        if avg_bars:
            recommendations.append(f"Avg time to target: <strong>{avg_bars:.0f} bars</strong>")

        if ar.expired > 0:
            recommendations.append(f"<span style='color:var(--red)'>{ar.expired} signals expired</span> — no target hit in window")

        if from_binance > 0 and from_excel > 0:
            recommendations.append(f"Data sources: <strong>{from_binance}</strong> live-verified, <strong>{from_excel}</strong> from Excel")

    # ---- Coin breakdown table ----
    coin_rows = ""
    for coin, stats in sorted(coin_stats.items(), key=lambda x: x[1]["wr"], reverse=True):
        wc = "wr-good" if stats["wr"] >= 55 else ("wr-bad" if stats["wr"] < 45 else "")
        coin_rows += f"""<tr>
            <td><strong>{coin}</strong></td>
            <td>{stats['count']}</td>
            <td class="{wc}">{stats['wr']}%</td>
            <td>{stats['won']}W / {stats['lost']}L</td>
        </tr>"""

    # ---- Daily summary ----
    daily_dates = sorted(ar.daily.keys())
    daily_rows = ""
    for day in daily_dates:
        trades = ar.daily[day]
        w = sum(1 for t in trades if t["won"])
        l = sum(1 for t in trades if t["lost"])
        be = len(trades) - w - l
        wr_val = _pct(w, w + l) if (w + l) > 0 else 0
        pnl_total = round(sum(t["pnl"] for t in trades), 2)
        daily_rows += f"""<tr>
            <td>{day}</td><td>{len(trades)}</td><td>{w+l}</td>
            <td>{w}</td><td>{l}</td><td>{be}</td>
            <td>{wr_val}%</td><td>{pnl_total:+.2f}%</td>
        </tr>"""

    # ---- Trade journal ----
    sorted_trades = sorted(all_trades, key=lambda t: t.timestamp or 0, reverse=True)
    journal_rows = ""
    for t in sorted_trades[:200]:
        pnl = t.pnl_pct
        pnl_str = f"{pnl:+.1f}%" if pnl is not None else "N/A"
        rr = t.achieved_rr
        rr_str = f"{rr:.1f}:1" if rr is not None else "N/A"
        ts = _parse_timestamp(t.timestamp)
        date_str = ts.strftime("%Y-%m-%d %H:%M") if ts else "N/A"
        row_class = "won" if t.result == "WON" else ("lost" if t.result == "LOST" else "expired")
        journal_rows += f"""
        <tr class="{row_class}">
            <td>{date_str}</td><td>{t.symbol}</td><td>{t.direction}</td>
            <td><span class="tier-badge tier-{t.tier.lower()}">{t.tier}</span></td>
            <td>{t.result}</td><td>{pnl_str}</td><td>{rr_str}</td>
            <td>{t.session}</td><td>{t.timeframe}</td>
            <td>{t.exit_reason or 'N/A'}</td>
        </tr>"""

    # ---- Colors ----
    wr_color = "#00b050" if win_rate >= 55 else ("#ff4444" if win_rate < 45 else "#ffcc00")
    pf_color = "#00b050" if pf >= 1.5 else ("#ff4444" if pf < 1.0 else "#ffcc00")

    # ---- CRT row helper ----
    def crt_row(label, condition, s):
        wr_val = f"{s['wr']}%" if s["count"] >= 3 else "—"
        return f'<tr><td>{label}</td><td>{condition}</td><td>{wr_val}</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'

    def first_s(d): return next(iter(d.values())) if d else {"wr": 0, "count": 0, "won": 0, "lost": 0}

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Judah Scanner — Signal Quality Audit</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {{ --bg: #0a0a1a; --card: #1a1a2e; --card-border: #2a2a4a; --accent: #00d4ff; --green: #00b050; --red: #ff4444; --yellow: #ffcc00; --text: #e0e0e0; --text-dim: #8888aa; }}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: var(--bg); color: var(--text); font-family: system-ui, -apple-system, 'Segoe UI', Roboto, sans-serif; line-height: 1.6; }}
.container {{ display: flex; max-width: 1600px; margin: 0 auto; }}
.sidebar {{ position: sticky; top: 0; width: 200px; height: 100vh; padding: 20px; border-right: 1px solid var(--card-border); overflow-y: auto; }}
.sidebar h3 {{ color: var(--accent); font-size: 0.7rem; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 10px; margin-top: 20px; }}
.sidebar a {{ display: block; color: var(--text-dim); text-decoration: none; font-size: 0.82rem; padding: 4px 0; transition: color 0.2s; }}
.sidebar a:hover {{ color: var(--accent); }}
.main {{ flex: 1; padding: 30px 40px; max-width: 1300px; }}
h1 {{ font-size: 1.6rem; margin-bottom: 5px; color: #fff; }}
.subtitle {{ color: var(--text-dim); font-size: 0.9rem; margin-bottom: 30px; }}
.section {{ margin-bottom: 40px; }}
.section-header {{ font-size: 0.75rem; text-transform: uppercase; letter-spacing: 2px; color: var(--accent); margin-bottom: 15px; padding-bottom: 8px; border-bottom: 1px solid var(--card-border); }}
.cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 25px; }}
.card {{ background: var(--card); border: 1px solid var(--card-border); border-radius: 12px; padding: 20px; text-align: center; }}
.card .label {{ font-size: 0.75rem; color: var(--text-dim); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }}
.card .value {{ font-size: 2rem; font-weight: 700; color: #fff; }}
.card .sub {{ font-size: 0.78rem; color: var(--text-dim); margin-top: 4px; }}
.chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
.chart-box {{ background: var(--card); border: 1px solid var(--card-border); border-radius: 12px; padding: 20px; }}
.chart-box h4 {{ font-size: 0.85rem; color: var(--text-dim); margin-bottom: 10px; text-transform: uppercase; letter-spacing: 1px; }}
.chart-box canvas {{ max-height: 300px; }}
.green-section {{ background: rgba(0,176,80,0.05); border: 1px solid rgba(0,176,80,0.2); border-radius: 12px; padding: 20px; }}
.red-section {{ background: rgba(255,68,68,0.05); border: 1px solid rgba(255,68,68,0.2); border-radius: 12px; padding: 20px; }}
.combo-item {{ display: flex; justify-content: space-between; align-items: center; padding: 8px 0; border-bottom: 1px solid var(--card-border); }}
.combo-item:last-child {{ border-bottom: none; }}
.combo-name {{ font-weight: 600; }}
.combo-stats {{ font-size: 0.85rem; color: var(--text-dim); text-align: right; }}
.combo-wr {{ font-weight: 700; color: var(--green); }}
table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
th {{ background: #0d0d20; color: var(--accent); padding: 10px 8px; text-align: left; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 1px; border-bottom: 2px solid var(--card-border); }}
td {{ padding: 8px; border-bottom: 1px solid var(--card-border); color: var(--text); }}
tr:hover td {{ background: rgba(255,255,255,0.02); }}
.wr-good {{ color: var(--green); font-weight: 600; }}
.wr-bad {{ color: var(--red); font-weight: 600; }}
.wr-mid {{ color: var(--yellow); font-weight: 600; }}
.won {{ background: rgba(0,176,80,0.08); }}
.lost {{ background: rgba(255,68,68,0.08); }}
.expired {{ background: rgba(255,204,0,0.05); }}
.tier-badge {{ padding: 2px 8px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; }}
.tier-sniper {{ background: rgba(0,212,255,0.15); color: var(--accent); }}
.tier-active {{ background: rgba(0,176,80,0.15); color: var(--green); }}
.tier-watch {{ background: rgba(255,204,0,0.15); color: var(--yellow); }}
.insufficient {{ background: rgba(255,204,0,0.08); border: 1px solid rgba(255,204,0,0.3); border-radius: 12px; padding: 15px 20px; color: var(--yellow); font-size: 0.9rem; }}
ul.recommendations {{ list-style: none; padding: 0; }}
ul.recommendations li {{ padding: 10px 0; border-bottom: 1px solid var(--card-border); font-size: 0.9rem; line-height: 1.5; }}
ul.recommendations li:last-child {{ border-bottom: none; }}
ul.recommendations li::before {{ content: "▸ "; color: var(--accent); font-weight: bold; }}
.two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
.method-badge {{ display: inline-block; padding: 3px 10px; border-radius: 4px; font-size: 0.75rem; font-weight: 600; margin-left: 8px; }}
.method-live {{ background: rgba(0,212,255,0.15); color: var(--accent); }}
.method-cached {{ background: rgba(255,204,0,0.15); color: var(--yellow); }}
@media (max-width: 900px) {{ .sidebar {{ display: none; }} .chart-grid, .two-col {{ grid-template-columns: 1fr; }} }}
.verify-info {{ background: rgba(0,212,255,0.05); border: 1px solid rgba(0,212,255,0.15); border-radius: 8px; padding: 12px 16px; margin-bottom: 20px; font-size: 0.85rem; color: var(--text-dim); }}
.verify-info strong {{ color: var(--accent); }}
.offline-warning {{ background: rgba(255,204,0,0.08); border: 1px solid rgba(255,204,0,0.3); border-radius: 8px; padding: 12px 16px; margin-bottom: 20px; color: var(--yellow); font-size: 0.85rem; }}
.offline-warning strong {{ color: var(--yellow); }}
</style>
</head>
<body>
<div class="container">
<nav class="sidebar">
    <h3>Navigation</h3>
    <a href="#summary">Executive Summary</a>
    <a href="#what-works">What Works</a>
    <a href="#what-avoid">What to Avoid</a>
    <a href="#breakdowns">Breakdowns</a>
    <a href="#coins">Per-Coin Results</a>
    <a href="#components">Components</a>
    <a href="#score">Score Analysis</a>
    <a href="#timeline">Timeline</a>
    <a href="#journal">Trade Journal</a>
    <a href="#recommendations">Recommendations</a>
</nav>
<div class="main">

<!-- HEADER -->
<div class="section">
    <h1>Judah Scanner — Signal Quality Audit Report</h1>
    <p class="subtitle">{date_range} &nbsp;|&nbsp; {ar.verified} verified trades <span class="{source_badge}">{source_label}</span></p>
    {('<div class="offline-warning"><strong>OFFLINE MODE:</strong> Binance API was unreachable. Outcomes loaded from Trade Outcomes sheet in Excel. For live verification, run without --offline flag.</div>') if offline else ''}
    <div class="verify-info">
        <strong>Verification:</strong> Each trade was verified by checking if TP was hit before SL.
        {f"Skipped: {verify_stats.get('skipped',0)} | Fetch failed: {verify_stats.get('fetch_failed',0)} | No TP/SL data: {verify_stats.get('no_data',0)}" if verify_stats else ''}
    </div>
</div>

<!-- SECTION 1: EXECUTIVE SUMMARY -->
<div class="section" id="summary">
    <div class="section-header">Executive Summary</div>
    {('<div class="insufficient">⚠️ ' + wr_text + '</div>') if insufficient else ''}
    <div class="cards">
        <div class="card">
            <div class="label">Win Rate</div>
            <div class="value" style="color:{wr_color}">{win_rate:.1f}%</div>
            <div class="sub">{ar.wins}W / {ar.losses}L / {ar.expired}EXP</div>
        </div>
        <div class="card">
            <div class="label">Avg RR</div>
            <div class="value">{avg_rr:.2f}:1</div>
            <div class="sub">target: 1.5:1</div>
        </div>
        <div class="card">
            <div class="label">Verified Trades</div>
            <div class="value">{total_closed}</div>
            <div class="sub">+ {ar.expired} expired</div>
        </div>
        <div class="card">
            <div class="label">Profit Factor</div>
            <div class="value" style="color:{pf_color}">{pf if pf < 999 else '∞'}</div>
            <div class="sub">gross profit / gross loss</div>
        </div>
        <div class="card">
            <div class="label">Avg PnL</div>
            <div class="value" style="color:{'var(--green)' if avg_pnl >= 0 else 'var(--red)'}">{avg_pnl:+.1f}%</div>
            <div class="sub">per trade average</div>
        </div>
        <div class="card">
            <div class="label">Avg Bars to Target</div>
            <div class="value">{avg_bars:.0f}</div>
            <div class="sub">candles to hit TP or SL</div>
        </div>
        <div class="card">
            <div class="label">Max Consec</div>
            <div class="value">{ar.max_consec_wins}W / {ar.max_consec_losses}L</div>
            <div class="sub">wins / losses streak</div>
        </div>
        <div class="card">
            <div class="label">Best / Worst</div>
            <div class="value" style="font-size:1.1rem;color:var(--green)">{best_str}</div>
            <div class="value" style="font-size:1.1rem;color:var(--red);margin-top:4px">{worst_str}</div>
        </div>
    </div>
</div>

<!-- SECTION 2: WHAT WORKS -->
<div class="section" id="what-works">
    <div class="section-header">What Works — Top 5 Combinations</div>
    <div class="green-section">
"""

    if top5:
        for c in top5:
            html += f"""        <div class="combo-item">
            <div><div class="combo-name">{c['combo']}</div></div>
            <div class="combo-stats"><span class="combo-wr">{c['wr']}% WR</span> &nbsp;({c['count']} trades, {c['won']}W/{c['lost']}L)</div>
        </div>"""
    else:
        html += '<p style="color:var(--text-dim)">Need 20+ closed trades for combination analysis.</p>'

    html += """
    </div>
</div>

<!-- SECTION 3: WHAT TO AVOID -->
<div class="section" id="what-avoid">
    <div class="section-header">What to Avoid — Bottom 3 Combinations</div>
    <div class="red-section">
"""

    if bottom3:
        for c in bottom3:
            html += f"""        <div class="combo-item">
            <div><div class="combo-name">{c['combo']}</div></div>
            <div class="combo-stats"><span style="color:var(--red);font-weight:700">{c['wr']}% WR</span> &nbsp;({c['count']} trades, {c['won']}W/{c['lost']}L)</div>
        </div>"""
    else:
        html += '<p style="color:var(--text-dim)">Need 20+ closed trades for combination analysis.</p>'

    html += """
    </div>
</div>

<!-- SECTION 4: BREAKDOWN CHARTS -->
<div class="section" id="breakdowns">
    <div class="section-header">Breakdown Charts</div>
    <div class="chart-grid">
        <div class="chart-box"><h4>Win Rate by Tier</h4><canvas id="chartTier"></canvas></div>
        <div class="chart-box"><h4>Win Rate by Session</h4><canvas id="chartSession"></canvas></div>
        <div class="chart-box"><h4>Win Rate by Timeframe</h4><canvas id="chartTF"></canvas></div>
        <div class="chart-box"><h4>Win Rate by Direction</h4><canvas id="chartDirection"></canvas></div>
    </div>
</div>

<!-- SECTION: PER-COIN RESULTS -->
<div class="section" id="coins">
    <div class="section-header">Per-Coin Performance</div>
    <div class="chart-box">
        <h4>Win Rate by Coin (sorted best to worst, min 3 trades)</h4>
        <table><tr><th>Coin</th><th>Trades</th><th>Win Rate</th><th>Won / Lost</th></tr>
"""
    if coin_rows:
        html += coin_rows
    else:
        html += '<tr><td colspan="4" style="color:var(--text-dim);text-align:center">Need 3+ trades per coin for this breakdown</td></tr>'
    html += """        </table>
    </div>
</div>

<!-- SECTION 5: COMPONENT ANALYSIS -->
<div class="section" id="components">
    <div class="section-header">Component Analysis</div>
    <div class="two-col">
        <div class="chart-box">
            <h4>CRT Components</h4>
            <table>
                <tr><th>Component</th><th>Condition</th><th>WR</th><th>Trades</th><th>W/L</th></tr>
"""
    html += crt_row("Displacement", ">= 2.0x avg body", first_s(disp_h))
    html += crt_row("Displacement", "< 2.0x avg body", first_s(disp_l))
    html += crt_row("OTE Zone", "In OTE (50-62%)", first_s(ote_i))
    html += crt_row("Optimal OTE", "50-62% retracement", first_s(opt_i))
    html += crt_row("Range Break", "Closed beyond range", first_s(rb_y))
    for zone_name in ["PREMIUM", "DISCOUNT", "EQUILIBRIUM"]:
        if zone_name in zone_stats and zone_stats[zone_name]["count"] >= 3:
            s = zone_stats[zone_name]
            html += f'<tr><td>Zone</td><td>{zone_name}</td><td>{s["wr"]}%</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'
    r618 = first_s(retr_stats.get("618_70", {})) if "618_70" in retr_stats else {"wr": 0, "count": 0}
    r50 = first_s(retr_stats.get("50_618", {})) if "50_618" in retr_stats else {"wr": 0, "count": 0}
    html += crt_row("Retracement", "61.8-70% (Golden pocket)", r618)
    html += crt_row("Retracement", "50-61.8% OTE", r50)

    html += """
            </table>
        </div>
        <div class="chart-box">
            <h4>SMC Components</h4>
            <table>
                <tr><th>Component</th><th>Condition</th><th>WR</th><th>Trades</th><th>W/L</th></tr>
"""
    html += crt_row("Order Block", "Has OB", first_s(ob_y))
    html += crt_row("Order Block", "No OB", first_s(ob_n))
    html += crt_row("FVG", "Has FVG", first_s(fvg_y))
    html += crt_row("MSB", "Has MSB", first_s(msb_y))
    html += crt_row("Liquidity Sweep", "Swept", first_s(liq_y))
    html += crt_row("Priority Boost", "Has boost", first_s(boost_y))
    for touches, s in sorted(ob_touch_stats.items()):
        if s["count"] >= 3:
            html += f'<tr><td>OB Retests</td><td>{touches}x touch</td><td>{s["wr"]}%</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'

    html += """
            </table>
        </div>
    </div>

    <!-- Confluence -->
    <div class="chart-box" style="margin-top:20px">
        <h4>Confluence Analysis</h4>
        <table>
            <tr><th>Confluence TFs</th><th>WR</th><th>Trades</th><th>Won / Lost</th></tr>
"""
    for k in ["0", "1", "2", "3+"]:
        if k in confluence_stats:
            s = confluence_stats[k]
            wc = "wr-good" if s["wr"] >= 55 else ("wr-bad" if s["wr"] < 45 else "")
            html += f'<tr><td>{k} timeframe(s)</td><td class="{wc}">{s["wr"]}%</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'

    html += """
        </table>
    </div>

    <!-- Freshness -->
    <div class="chart-box" style="margin-top:20px">
        <h4>Freshness Analysis</h4>
        <table>
            <tr><th>Freshness at Entry</th><th>WR</th><th>Trades</th><th>Won / Lost</th></tr>
"""
    for k in ["hot", "warm", "cool", "cold", "dead", "expired"]:
        if k in freshness_stats:
            s = freshness_stats[k]
            html += f'<tr><td>{k.title()}</td><td>{s["wr"]}%</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'

    html += """
        </table>
    </div>

    <!-- Bars -->
    <div class="chart-box" style="margin-top:20px">
        <h4>Time to Target</h4>
        <table>
            <tr><th>Bars to Exit</th><th>WR</th><th>Trades</th><th>Won / Lost</th></tr>
"""
    for k in ["1-2 bars", "3-8 bars", "9-24 bars", "25+ bars"]:
        if k in bars_stats:
            s = bars_stats[k]
            html += f'<tr><td>{k}</td><td>{s["wr"]}%</td><td>{s["count"]}</td><td>{s["won"]}W/{s["lost"]}L</td></tr>'

    html += """
        </table>
    </div>
</div>

<!-- SECTION 6: SCORE ANALYSIS -->
<div class="section" id="score">
    <div class="section-header">Score Analysis</div>
    <div class="chart-box">
        <h4>Win Rate by Composite Score Range</h4>
        <canvas id="chartScore"></canvas>
    </div>
</div>

<!-- SECTION 7: TIMELINE -->
<div class="section" id="timeline">
    <div class="section-header">Performance Timeline</div>
    <div class="chart-grid">
        <div class="chart-box"><h4>Equity Curve (Cumulative PnL %)</h4><canvas id="chartEquity"></canvas></div>
"""

    if daily_dates:
        daily_wr = []
        daily_pnl = []
        for day in daily_dates:
            trades = ar.daily[day]
            w = sum(1 for t in trades if t["won"])
            l = sum(1 for t in trades if t["lost"])
            wr_val = round(w / (w + l) * 100, 1) if (w + l) > 0 else 0
            daily_wr.append(wr_val)
            daily_pnl.append(round(sum(t["pnl"] for t in trades), 2))
        html += """
        <div class="chart-box"><h4>Daily Win Rate & PnL</h4><canvas id="chartDaily"></canvas></div>
"""

    html += """
    </div>
</div>

<!-- SECTION 8: TRADE JOURNAL -->
<div class="section" id="journal">
    <div class="section-header">Trade Journal (all verified trades)</div>
    <div class="chart-box" style="overflow-x:auto">
        <table>
            <tr>
                <th>Close Time</th><th>Symbol</th><th>Direction</th><th>Tier</th>
                <th>Result</th><th>PnL %</th><th>RR</th><th>Session</th>
                <th>TF</th><th>Exit Reason</th>
            </tr>
"""
    html += journal_rows
    html += """
        </table>
    </div>
</div>

<!-- SECTION 9: RECOMMENDATIONS -->
<div class="section" id="recommendations">
    <div class="section-header">Recommendations</div>
    <div class="green-section">
        <ul class="recommendations">
"""
    for rec in recommendations:
        html += f"<li>{rec}</li>"
    if not recommendations:
        html += "<li>Collect more trade data for personalized recommendations.</li>"
    html += """
        </ul>
    </div>
</div>

</div>
</div>

<script>
Chart.defaults.color = '#8888aa';
Chart.defaults.borderColor = '#2a2a4a';
Chart.defaults.font.family = "system-ui, -apple-system, sans-serif";

const CD = {
    responsive: true, maintainAspectRatio: true,
    plugins: { legend: { labels: { color: '#e0e0e0', padding: 15 } } },
    scales: {
        x: { ticks: { color: '#8888aa' }, grid: { color: '#1a1a2e' } },
        y: { ticks: { color: '#8888aa' }, grid: { color: '#2a2a4a' } }
    }
};

"""

    if tier_stats:
        labels = json.dumps(list(tier_stats.keys()))
        values = json.dumps([tier_stats[k]["wr"] for k in tier_stats])
        counts = json.dumps([tier_stats[k]["count"] for k in tier_stats])
        html += f"""
new Chart(document.getElementById('chartTier'), {{
    type: 'bar',
    data: {{ labels: {labels}, datasets: [{{
        label: 'Win Rate %', data: {values},
        backgroundColor: ['#00d4ff','#00b050','#ffcc00'], borderColor: ['#00d4ff','#00b050','#ffcc00'],
        borderWidth: 1, borderRadius: 6,
    }}] }},
    options: {{ ...CD, plugins: {{ ...CD.plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + '% WR (' + {counts}[ctx.dataIndex] + ' trades)' }} }} }},
        scales: {{ ...CD.scales, y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }} }} }}
}});
"""

    if session_stats:
        labels = json.dumps(list(session_stats.keys()))
        values = json.dumps([session_stats[k]["wr"] for k in session_stats])
        counts = json.dumps([session_stats[k]["count"] for k in session_stats])
        html += f"""
new Chart(document.getElementById('chartSession'), {{
    type: 'bar',
    data: {{ labels: {labels}, datasets: [{{
        label: 'Win Rate %', data: {values},
        backgroundColor: '#00d4ff', borderColor: '#00d4ff',
        borderWidth: 1, borderRadius: 6,
    }}] }},
    options: {{ ...CD, plugins: {{ ...CD.plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + '% WR (' + {counts}[ctx.dataIndex] + ' trades)' }} }} }},
        scales: {{ ...CD.scales, y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }} }} }}
}});
"""

    if tf_stats:
        labels = json.dumps(list(tf_stats.keys()))
        values = json.dumps([tf_stats[k]["wr"] for k in tf_stats])
        counts = json.dumps([tf_stats[k]["count"] for k in tf_stats])
        html += f"""
new Chart(document.getElementById('chartTF'), {{
    type: 'bar',
    data: {{ labels: {labels}, datasets: [{{
        label: 'Win Rate %', data: {values},
        backgroundColor: '#00b050', borderColor: '#00b050',
        borderWidth: 1, borderRadius: 6,
    }}] }},
    options: {{ ...CD, plugins: {{ ...CD.plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + '% WR (' + {counts}[ctx.dataIndex] + ' trades)' }} }} }},
        scales: {{ ...CD.scales, y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }} }} }}
}});
"""

    if direction_stats:
        labels = json.dumps(list(direction_stats.keys()))
        values = json.dumps([direction_stats[k]["wr"] for k in direction_stats])
        counts = json.dumps([direction_stats[k]["count"] for k in direction_stats])
        colors = json.dumps(["#00d4ff" if k.upper() == "BULLISH" else "#ff4444" for k in direction_stats])
        html += f"""
new Chart(document.getElementById('chartDirection'), {{
    type: 'bar',
    data: {{ labels: {labels}, datasets: [{{
        label: 'Win Rate %', data: {values},
        backgroundColor: {colors}, borderColor: {colors},
        borderWidth: 1, borderRadius: 6,
    }}] }},
    options: {{ ...CD, plugins: {{ ...CD.plugins, tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + '% WR (' + {counts}[ctx.dataIndex] + ' trades)' }} }} }},
        scales: {{ ...CD.scales, y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }} }} }}
}});
"""

    score_order = ["<30", "30-39", "40-49", "50-59", "60-69", "70-79", "80-100"]
    score_labels = [s for s in score_order if s in score_stats]
    if score_labels:
        score_values = [score_stats[s]["wr"] for s in score_labels]
        score_counts = [score_stats[s]["count"] for s in score_labels]
        bar_colors = json.dumps(["#00b050" if v >= 55 else ("#ff4444" if v < 45 else "#ffcc00") for v in score_values])
        html += f"""
new Chart(document.getElementById('chartScore'), {{
    type: 'bar',
    data: {{
        labels: {json.dumps(score_labels)},
        datasets: [{{
            label: 'Win Rate %', data: {json.dumps(score_values)},
            backgroundColor: {bar_colors}, borderWidth: 1, borderRadius: 6,
        }}, {{
            label: 'Trade Count', data: {json.dumps(score_counts)},
            backgroundColor: 'rgba(0,212,255,0.2)', borderColor: '#00d4ff',
            borderWidth: 1, borderRadius: 6, yAxisID: 'y1',
        }}]
    }},
    options: {{ ...CD,
        scales: {{ ...CD.scales,
            y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }},
            y1: {{ position: 'right', grid: {{ drawOnChartArea: false }}, ticks: {{ color: '#00d4ff' }} }}
        }}
    }}
}});
"""

    if ar.equity_curve:
        eq_dates = json.dumps([e["date"] for e in ar.equity_curve])
        eq_vals = json.dumps([e["cumulative"] for e in ar.equity_curve])
        html += f"""
new Chart(document.getElementById('chartEquity'), {{
    type: 'line',
    data: {{ labels: {eq_dates}, datasets: [{{
        label: 'Cumulative PnL %', data: {eq_vals},
        borderColor: '#00d4ff', backgroundColor: 'rgba(0,212,255,0.1)',
        fill: true, tension: 0.3, pointRadius: 2,
    }}] }},
    options: CD
}});
"""

    if daily_dates:
        html += f"""
new Chart(document.getElementById('chartDaily'), {{
    type: 'line',
    data: {{
        labels: {json.dumps(daily_dates)},
        datasets: [{{
            label: 'Win Rate %', data: {json.dumps(daily_wr)},
            borderColor: '#00b050', backgroundColor: 'rgba(0,176,80,0.1)',
            fill: true, tension: 0.3,
        }}, {{
            label: 'PnL %', data: {json.dumps(daily_pnl)},
            borderColor: '#ffcc00', backgroundColor: 'rgba(255,204,0,0.1)',
            fill: true, tension: 0.3, yAxisID: 'y1',
        }}]
    }},
    options: {{ ...CD,
        scales: {{ ...CD.scales,
            y: {{ ...CD.scales.y, min: 0, max: 100, ticks: {{ ...CD.scales.y.ticks, callback: v => v + '%' }} }},
            y1: {{ position: 'right', grid: {{ drawOnChartArea: false }}, ticks: {{ color: '#ffcc00', callback: v => v + '%' }} }}
        }}
    }}
}});
"""

    html += """
</script>
</body>
</html>"""

    return html


# ============================================================
# MAIN
# ============================================================

def main():
    offline = "--offline" in sys.argv

    filepath = os.path.join(os.getcwd(), "signal_log.xlsx")
    output_path = os.path.join(os.getcwd(), "quality_report.html")

    print("=" * 60)
    print("  Judah Scanner — Signal Quality Auditor")
    if offline:
        print("  MODE: OFFLINE (Excel data only)")
    else:
        print("  MODE: LIVE VERIFICATION (Binance API)")
    print("=" * 60)

    wb = load_workbook(filepath)
    signals = load_signals_from_excel(wb)
    wb.close()

    print(f"\nLoaded {len(signals)} signals from Excel")

    verified_trades, verify_stats = verify_signals(signals, offline=offline)

    if not verified_trades:
        print("\nERROR: No trades could be analyzed.")
        print("  - If using --offline: add Trade Outcomes to signal_log.xlsx")
        print("  - If online: check internet connectivity")
        sys.exit(1)

    print(f"\nAnalyzing {len(verified_trades)} verified trades...")
    ar = analyze(verified_trades)

    html = generate_html(ar, verified_trades, verify_stats)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    total_closed = ar.wins + ar.losses
    win_rate = _pct(ar.wins, total_closed) if total_closed > 0 else 0.0
    pf = round(ar.gross_profit / ar.gross_loss, 2) if ar.gross_loss > 0 else 0.0

    print(f"\nDone!")
    print(f"  Wins: {ar.wins} | Losses: {ar.losses} | Expired: {ar.expired}")
    print(f"  Win Rate: {win_rate:.1f}%")
    print(f"  Profit Factor: {pf}")
    print(f"  Avg PnL: {_avg(ar.all_pnl):+.1f}%")
    print(f"  Report: {output_path}")


if __name__ == "__main__":
    main()

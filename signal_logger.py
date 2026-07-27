#!/usr/bin/env python3
"""
signal_logger.py — Real-time Excel logger for Judah Scanner signals.

Connects to the Judah Scanner WebSocket server and writes signals to an
Excel workbook with 4 sheets in real-time.

Usage:
    python signal_logger.py [--ws-url ws://localhost:8000] [--output signal_log.xlsx]

Requirements:
    pip install openpyxl websockets
"""

import argparse
import json
import logging
import os
import sys
import threading
import time
from datetime import datetime, timezone
from queue import Queue, Empty

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import websockets
except ImportError:
    print("ERROR: websockets not installed. Run: pip install websockets")
    sys.exit(1)

try:
    import aiohttp
except ImportError:
    aiohttp = None
    print("[warn] aiohttp not installed — REST API fallback disabled. Run: pip install aiohttp")

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("signal_logger")

# ─── Column Definitions ───────────────────────────────────────────────────────

HEADERS_SIGNALS = [
    "Timestamp", "Signal ID", "Symbol", "Timeframe", "Engine",
    "Direction", "Tier", "Composite Score", "Base Score",
    "CRT Score", "SMC Score",
    "Confluence Count", "Confluence Details", "Confluence Boost",
    "Priority Boosts",
    "Entry Price", "Stop Loss", "Take Profit",
    "RR Ratio", "Risk ($)", "Reward ($)",
    "Current Price", "Distance to Entry %",
    "Session", "Session Score", "Scenario",
    "Freshness State", "Freshness Factor", "Age Minutes",
    "Score at Birth", "Live Score", "Decay Applied",
    "RR Quality", "Status",
]

HEADERS_CRT = [
    "Signal ID", "Timestamp", "Symbol",
    "Displacement Ratio", "Displacement Direction", "Displacement High",
    "Displacement Low", "Displacement Is Extreme",
    "Retracement %", "In OTE", "In Optimal OTE",
    "Range Low", "Range High", "Range Midpoint", "Range Size",
    "Range Break", "Session", "Session Score",
    "ATR Value", "ATR %",
    "Premium/Discount", "Price Position %",
]

HEADERS_SMC = [
    "Signal ID", "Timestamp", "Symbol",
    "OB Type", "OB Low", "OB High", "OB Retest Count",
    "FVG Type", "FVG Bottom", "FVG Top", "FVG Filled",
    "MSB Confirmed", "MSB Type", "MSB Direction", "MSB Level",
    "Liquidity Swept", "Liquidity Level", "Liquidity Direction",
    "VSP Type", "VSP Price", "VSP Body Quality %",
    "Confluence Count", "Confluence Details",
]

HEADERS_TRADE = [
    "Signal ID", "Timestamp", "Symbol", "Direction",
    "Entry Price", "Initial SL", "Initial TP", "Initial RR",
    "Risk $", "Reward $",
    "Current Price", "Distance to Entry %",
    "Current RR", "Status", "Closed At Price", "Closed PnL $", "Closed PnL %",
]

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _v(sig, key, default="N/A"):
    """Safe dict get — handles None keys, missing keys, and 'N/A' default."""
    if not sig or not isinstance(sig, dict):
        return default
    val = sig.get(key)
    if val is None:
        return default
    return val

def _join(arr):
    """Join a list into a comma-separated string, or return 'N/A'."""
    if not arr:
        return "N/A"
    if isinstance(arr, list):
        return ", ".join(str(x) for x in arr)
    return str(arr)

def _rr_quality(rr):
    """Classify RR ratio into Good/OK/Poor."""
    try:
        r = float(rr)
        if r >= 2.0:
            return "Good"
        if r >= 1.5:
            return "OK"
        return "Poor"
    except (TypeError, ValueError):
        return "N/A"

def _yes_no(val):
    """Convert truthy/falsy to Yes/No."""
    if val is None:
        return "No"
    return "Yes" if val else "No"

def _fmt_number(val, decimals=2):
    """Format a number, return 'N/A' if not numeric."""
    if val is None:
        return "N/A"
    try:
        return str(round(float(val), decimals))
    except (TypeError, ValueError):
        return "N/A"

def _fmt_price(val):
    """Format price with smart decimals."""
    if val is None:
        return "N/A"
    try:
        f = float(val)
        if f >= 1000:
            return f"${f:,.2f}"
        if f >= 1:
            return f"${f:.3f}"
        return f"${f:.5f}"
    except (TypeError, ValueError):
        return "N/A"

# ─── Row Extraction ───────────────────────────────────────────────────────────

def extract_signal_row(sig: dict) -> list:
    """Extract a row for the Signals sheet."""
    row = [
        _v(sig, "timestamp"),           # Timestamp
        _v(sig, "id"),                   # Signal ID
        _v(sig, "symbol"),              # Symbol
        _v(sig, "timeframe"),           # Timeframe
        _v(sig, "engine"),              # Engine
        _v(sig, "direction"),           # Direction
        _v(sig, "tier"),                # Tier
        _v(sig, "composite_score") or _v(sig, "compositeScore"),  # Composite Score
        _v(sig, "base_score"),          # Base Score
        _v(sig, "crt_score") or _v(sig, "crtScore"),              # CRT Score
        _v(sig, "smc_score") or _v(sig, "smcScore"),              # SMC Score
        len(_v(sig, "confluence", [])), # Confluence Count
        _join(_v(sig, "confluence", [])), # Confluence Details
        _v(sig, "confluence_boost"),    # Confluence Boost
        _join(_v(sig, "priority_boosts")), # Priority Boosts
        _fmt_price(_v(sig, "entry")),   # Entry Price
        _fmt_price(_v(sig, "stop_loss")),  # Stop Loss
        _fmt_price(_v(sig, "take_profit")), # Take Profit
        _v(sig, "rr"),                  # RR Ratio
        _fmt_number(_v(sig, "risk"), 2),  # Risk $
        _fmt_number(_v(sig, "reward"), 2), # Reward $
        _fmt_price(_v(sig, "current_price")), # Current Price
        _fmt_number(_v(sig, "distance_to_entry_pct"), 2), # Distance to Entry %
        _v(sig, "session"),             # Session
        _v(sig, "session_score"),       # Session Score
        _v(sig, "scenario"),            # Scenario
        _v(sig, "freshness_state"),     # Freshness State
        _v(sig, "freshness_factor"),    # Freshness Factor
        _v(sig, "age_minutes"),         # Age Minutes
        _v(sig, "base_score"),          # Score at Birth (same as base_score at creation)
        _v(sig, "composite_score") or _v(sig, "compositeScore"),  # Live Score
        _v(sig, "decay"),               # Decay Applied
        _rr_quality(_v(sig, "rr")),     # RR Quality
        _v(sig, "status", "ACTIVE"),    # Status
    ]
    return row


def extract_crt_row(sig: dict) -> list:
    """Extract a row for the CRT Analysis sheet."""
    disp = _v(sig, "displacement") or {}
    rng = _v(sig, "range") or {}

    row = [
        _v(sig, "id"),                  # Signal ID
        _v(sig, "timestamp"),           # Timestamp
        _v(sig, "symbol"),              # Symbol
        f"{_v(disp, 'ratio', 'N/A')}x", # Displacement Ratio
        _v(disp, "direction"),          # Displacement Direction
        _fmt_number(_v(disp, "high"), 2),  # Displacement High
        _fmt_number(_v(disp, "low"), 2),   # Displacement Low
        _yes_no(_v(disp, "is_extreme")),  # Displacement Is Extreme
        _v(sig, "retracement_percent"), # Retracement %
        _yes_no(_v(sig, "in_ote")),     # In OTE
        _yes_no(_v(sig, "in_optimal_ote")), # In Optimal OTE
        _fmt_number(_v(rng, "low"), 2),   # Range Low
        _fmt_number(_v(rng, "high"), 2),  # Range High
        _fmt_number(_v(rng, "midpoint"), 2), # Range Midpoint
        _fmt_number(_v(rng, "range_size"), 2), # Range Size
        _v(sig, "range_break"),         # Range Break
        _v(sig, "session"),             # Session
        _v(sig, "session_score"),       # Session Score
        _fmt_number(_v(sig, "atr_value"), 4), # ATR Value
        _fmt_number(_v(sig, "atr_percent"), 2), # ATR %
        _v(sig, "premium_discount"),    # Premium/Discount Zone
        _v(sig, "price_position_pct"),  # Price Position %
    ]
    return row


def extract_smc_row(sig: dict) -> list:
    """Extract a row for the SMC Analysis sheet."""
    ob = _v(sig, "ob") or {}
    fvg = _v(sig, "fvg") or {}
    msb = _v(sig, "msb") or {}
    liq = _v(sig, "liquidity") or {}
    vsp = _v(sig, "vsp") or {}

    row = [
        _v(sig, "id"),                  # Signal ID
        _v(sig, "timestamp"),           # Timestamp
        _v(sig, "symbol"),              # Symbol
        _v(ob, "type"),                 # OB Type
        _fmt_number(_v(ob, "low"), 2),  # OB Low
        _fmt_number(_v(ob, "high"), 2), # OB High
        _v(ob, "touches"),              # OB Retest Count
        _v(fvg, "type"),                # FVG Type
        _fmt_number(_v(fvg, "bottom"), 2),  # FVG Bottom
        _fmt_number(_v(fvg, "top"), 2),     # FVG Top
        _yes_no(_v(fvg, "filled")),     # FVG Filled
        _yes_no(_v(msb, "confirmed")),  # MSB Confirmed
        _v(msb, "type"),                # MSB Type
        _v(msb, "direction"),           # MSB Direction
        _fmt_number(_v(msb, "level"), 2),   # MSB Level
        _yes_no(_v(liq, "swept")),      # Liquidity Swept
        _fmt_number(_v(liq, "level"), 2),   # Liquidity Level
        _v(liq, "direction"),           # Liquidity Direction
        _v(vsp, "type"),                # VSP Type
        _fmt_number(_v(vsp, "price"), 2),   # VSP Price
        _fmt_number(_v(vsp, "body_quality"), 1),  # VSP Body Quality %
        len(_v(sig, "confluence", [])), # Confluence Count
        _join(_v(sig, "confluence", [])), # Confluence Details
    ]
    return row


def extract_trade_row(sig: dict) -> list:
    """Extract a row for the Trade Levels sheet."""
    row = [
        _v(sig, "id"),                  # Signal ID
        _v(sig, "timestamp"),           # Timestamp
        _v(sig, "symbol"),              # Symbol
        _v(sig, "direction"),           # Direction
        _fmt_price(_v(sig, "entry")),   # Entry Price
        _fmt_price(_v(sig, "stop_loss")), # Initial SL
        _fmt_price(_v(sig, "take_profit")), # Initial TP
        _v(sig, "rr"),                  # Initial RR
        _fmt_number(_v(sig, "risk"), 2),  # Risk $
        _fmt_number(_v(sig, "reward"), 2), # Reward $
        _fmt_price(_v(sig, "current_price")), # Current Price
        _fmt_number(_v(sig, "distance_to_entry_pct"), 2), # Distance to Entry %
        _v(sig, "rr"),                  # Current RR
        _v(sig, "status", "PENDING"),   # Status
        "N/A",                           # Closed At Price
        "N/A",                           # Closed PnL $
        "N/A",                           # Closed PnL %
    ]
    return row

# ─── Workbook Manager ─────────────────────────────────────────────────────────

class WorkbookManager:
    """Manages the Excel workbook with 4 sheets, formatting, and thread safety."""

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._lock = threading.Lock()
        self._wb = None
        self._sheets = {}
        self._signal_row_map = {}   # signal_id -> row_num in Signals sheet
        self._signal_trade_map = {} # signal_id -> row_num in Trade Levels sheet

    def create(self, signals: list):
        """Create a fresh workbook from INITIAL signals."""
        with self._lock:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            ws_signals = wb.create_sheet("Signals")
            ws_crt = wb.create_sheet("CRT Analysis")
            ws_smc = wb.create_sheet("SMC Analysis")
            ws_trade = wb.create_sheet("Trade Levels")

            # Write headers
            for ws, headers in [
                (ws_signals, HEADERS_SIGNALS),
                (ws_crt, HEADERS_CRT),
                (ws_smc, HEADERS_SMC),
                (ws_trade, HEADERS_TRADE),
            ]:
                ws.append(headers)
                self._format_header_row(ws, len(headers))

            # Write data rows
            row_map = {}
            trade_map = {}
            for sig in signals:
                sig_id = _v(sig, "id")
                row_num = ws_signals.max_row + 1
                ws_signals.append(extract_signal_row(sig))
                row_map[sig_id] = row_num

                ws_crt.append(extract_crt_row(sig))
                ws_smc.append(extract_smc_row(sig))

                trade_row_num = ws_trade.max_row + 1
                ws_trade.append(extract_trade_row(sig))
                trade_map[sig_id] = trade_row_num

                # Alternate row coloring
                if row_num % 2 == 0:
                    for ws in [ws_signals, ws_crt, ws_smc, ws_trade]:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row_num, column=col).fill = self.ALT_FILL

            # Apply conditional formatting
            self._apply_conditional_formatting(ws_signals)

            # Auto-size columns
            for ws in [ws_signals, ws_crt, ws_smc, ws_trade]:
                self._auto_size_columns(ws)

            # Freeze header rows
            for ws in [ws_signals, ws_crt, ws_smc, ws_trade]:
                ws.freeze_panes = "A2"

            # Save
            self._save_with_retry(wb)
            self._wb = wb
            self._sheets = {
                "signals": ws_signals,
                "crt": ws_crt,
                "smc": ws_smc,
                "trade": ws_trade,
            }
            self._signal_row_map = row_map
            self._signal_trade_map = trade_map
            logger.info(f"[excel] Created workbook with {len(signals)} signals across 4 sheets")

    def add_signals(self, signals: list):
        """Append new signals to the workbook."""
        if not signals:
            return
        with self._lock:
            self._reload()
            ws_s = self._sheets["signals"]
            ws_c = self._sheets["crt"]
            ws_m = self._sheets["smc"]
            ws_t = self._sheets["trade"]

            for sig in signals:
                sig_id = _v(sig, "id")
                row_num = ws_s.max_row + 1

                ws_s.append(extract_signal_row(sig))
                ws_c.append(extract_crt_row(sig))
                ws_m.append(extract_smc_row(sig))

                trade_row_num = ws_t.max_row + 1
                ws_t.append(extract_trade_row(sig))

                self._signal_row_map[sig_id] = row_num
                self._signal_trade_map[sig_id] = trade_row_num

                # Alternate row coloring
                if row_num % 2 == 0:
                    for ws in [ws_s, ws_c, ws_m, ws_t]:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row_num, column=col).fill = self.ALT_FILL

            self._save_with_retry(self._wb)
            logger.info(f"[excel] Added {len(signals)} new signals")

    def refresh_signals(self, signals: list):
        """Update existing signal rows with refreshed data (current price, score, etc.)."""
        if not signals:
            return
        with self._lock:
            self._reload()
            ws_s = self._sheets["signals"]
            ws_t = self._sheets["trade"]

            updated = 0
            for sig in signals:
                sig_id = _v(sig, "id")
                row_num = self._signal_row_map.get(sig_id)
                trade_row = self._signal_trade_map.get(sig_id)

                if row_num and row_num <= ws_s.max_row:
                    # Col indices: CurrentPrice=22, Dist%=23, Freshness=27, Age=28
                    # ScoreBirth=29, LiveScore=30, Decay=31, RRQuality=32, Status=33
                    self._safe_set(ws_s, row_num, 22, _fmt_price(_v(sig, "current_price")))
                    self._safe_set(ws_s, row_num, 23, _fmt_number(_v(sig, "distance_to_entry_pct"), 2))
                    self._safe_set(ws_s, row_num, 30, _v(sig, "composite_score") or _v(sig, "compositeScore"))
                    self._safe_set(ws_s, row_num, 27, _v(sig, "freshness_state"))
                    self._safe_set(ws_s, row_num, 28, _v(sig, "age_minutes"))

                    # Recalculate decay: score_at_birth - live_score
                    base = _v(sig, "base_score", 0)
                    try:
                        live = float(_v(sig, "composite_score") or 0)
                        base_f = float(base)
                        decay = max(0, round(base_f - live, 1))
                    except (TypeError, ValueError):
                        decay = _v(sig, "decay", 0)
                    self._safe_set(ws_s, row_num, 31, decay)

                    updated += 1

                if trade_row and trade_row <= ws_t.max_row:
                    self._safe_set(ws_t, trade_row, 11, _fmt_price(_v(sig, "current_price")))
                    self._safe_set(ws_t, trade_row, 12, _fmt_number(_v(sig, "distance_to_entry_pct"), 2))

            if updated > 0:
                self._save_with_retry(self._wb)
                logger.info(f"[excel] Refreshed {updated} signals")

    def expire_signals(self, signals: list):
        """Mark signals as EXPIRED in both Signals and Trade Levels sheets."""
        if not signals:
            return
        with self._lock:
            self._reload()
            ws_s = self._sheets["signals"]
            ws_t = self._sheets["trade"]

            updated = 0
            for sig in signals:
                sig_id = _v(sig, "id")
                row_num = self._signal_row_map.get(sig_id)
                trade_row = self._signal_trade_map.get(sig_id)

                # Status = col 34 in Signals, col 14 in Trade
                if row_num and row_num <= ws_s.max_row:
                    self._safe_set(ws_s, row_num, 34, "EXPIRED")
                    updated += 1

                if trade_row and trade_row <= ws_t.max_row:
                    self._safe_set(ws_t, trade_row, 14, "EXPIRED")

            if updated > 0:
                self._save_with_retry(self._wb)
                logger.info(f"[excel] Marked {updated} signals as EXPIRED")

    def _reload(self):
        """Reload workbook from disk if it exists (handles in-place updates)."""
        if os.path.exists(self.filepath):
            try:
                self._wb = openpyxl.load_workbook(self.filepath)
                self._sheets = {
                    "signals": self._wb["Signals"],
                    "crt": self._wb["CRT Analysis"],
                    "smc": self._wb["SMC Analysis"],
                    "trade": self._wb["Trade Levels"],
                }
                # Rebuild index after reload — row numbers may have changed
                self._rebuild_index()
                return
            except Exception as e:
                logger.warning(f"[excel] Reload failed: {e}, creating new workbook")

        # Create fresh workbook
        self._wb = Workbook()
        self._wb.remove(self._wb.active)
        for name, headers in [
            ("Signals", HEADERS_SIGNALS),
            ("CRT Analysis", HEADERS_CRT),
            ("SMC Analysis", HEADERS_SMC),
            ("Trade Levels", HEADERS_TRADE),
        ]:
            ws = self._wb.create_sheet(name)
            ws.append(headers)
            self._format_header_row(ws, len(headers))
        self._sheets = {
            "signals": self._wb["Signals"],
            "crt": self._wb["CRT Analysis"],
            "smc": self._wb["SMC Analysis"],
            "trade": self._wb["Trade Levels"],
        }
        self._signal_row_map.clear()
        self._signal_trade_map.clear()

    def _safe_set(self, ws, row, col, value):
        """Set a cell value, handling type conversions."""
        try:
            # Try numeric conversion for cleaner Excel
            if isinstance(value, str) and value not in ("N/A", ""):
                try:
                    if "." in value:
                        value = float(value)
                    else:
                        value = int(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row, column=col).value = value
        except Exception as e:
            logger.debug(f"[excel] Set error r{row}c{col}: {e}")

    def _save_with_retry(self, wb, max_retries=3):
        """Save workbook with retry logic for when file is open in Excel."""
        for attempt in range(max_retries):
            try:
                wb.save(self.filepath)
                return
            except PermissionError:
                if attempt < max_retries - 1:
                    logger.warning(
                        f"[excel] File is locked (open in Excel?). "
                        f"Retrying in 5s... ({attempt + 1}/{max_retries})"
                    )
                    time.sleep(5)
                else:
                    logger.error(
                        f"[excel] Cannot save — file '{self.filepath}' is locked. "
                        "Close the Excel file and the script will catch up on next write."
                    )
            except Exception as e:
                logger.error(f"[excel] Save error: {e}")
                return

    def _format_header_row(self, ws, num_cols):
        """Apply header formatting."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.THIN_BORDER

    def _apply_conditional_formatting(self, ws):
        """Apply color scales and rules to the Signals sheet."""
        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

        # Composite Score column (H = col 8) — green/yellow/red
        score_col = "H"
        ws.conditional_formatting.add(
            f"{score_col}2:{score_col}1048576",
            CellIsRule(
                operator="greaterThan", formula=["60"],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                font=Font(color="006100"),
            ),
        )
        ws.conditional_formatting.add(
            f"{score_col}2:{score_col}1048576",
            CellIsRule(
                operator="between", formula=["40", "60"],
                fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                font=Font(color="9C5700"),
            ),
        )
        ws.conditional_formatting.add(
            f"{score_col}2:{score_col}1048576",
            CellIsRule(
                operator="lessThan", formula=["40"],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006"),
            ),
        )

        # Freshness State column (AA = col 27)
        freshness_col = "AA"
        freshness_colors = {
            "hot": "FF6B6B",
            "warm": "FFA94D",
            "cool": "FFD43B",
            "cold": "74C0FC",
            "dead": "868E96",
        }
        for state, color in freshness_colors.items():
            ws.conditional_formatting.add(
                f"{freshness_col}2:{freshness_col}1048576",
                CellIsRule(
                    operator="equal", formula=[f'"{state}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                    font=Font(color="FFFFFF" if state in ("hot", "dead") else "000000"),
                ),
            )

        # RR Quality column (AG = col 33)
        rr_col = "AG"
        ws.conditional_formatting.add(
            f"{rr_col}2:{rr_col}1048576",
            CellIsRule(
                operator="equal", formula=['"Good"'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                font=Font(color="006100"),
            ),
        )
        ws.conditional_formatting.add(
            f"{rr_col}2:{rr_col}1048576",
            CellIsRule(
                operator="equal", formula=['"OK"'],
                fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                font=Font(color="9C5700"),
            ),
        )
        ws.conditional_formatting.add(
            f"{rr_col}2:{rr_col}1048576",
            CellIsRule(
                operator="equal", formula=['"Poor"'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006"),
            ),
        )

    def _auto_size_columns(self, ws):
        """Auto-size columns based on content."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            adjusted = min(max_len + 3, 50)  # Cap at 50 chars wide
            ws.column_dimensions[col_letter].width = max(adjusted, 8)

    def rebuild_index(self):
        """Rebuild the signal_id -> row_num index from the current workbook."""
        if not self._wb or not os.path.exists(self.filepath):
            return
        try:
            ws_s = self._wb["Signals"]
            ws_t = self._wb["Trade Levels"]
            self._signal_row_map.clear()
            self._signal_trade_map.clear()

            # Signal ID is column B (col 2)
            for row in range(2, ws_s.max_row + 1):
                sig_id = ws_s.cell(row=row, column=2).value
                if sig_id:
                    self._signal_row_map[sig_id] = row

            for row in range(2, ws_t.max_row + 1):
                sig_id = ws_t.cell(row=row, column=1).value
                if sig_id:
                    self._signal_trade_map[sig_id] = row
        except Exception as e:
            logger.debug(f"[excel] Index rebuild: {e}")

# ─── WebSocket Client ─────────────────────────────────────────────────────────

class SignalLogger:
    """Main orchestrator: WebSocket client + Excel writer."""

    def __init__(self, ws_url: str, output_file: str):
        self.ws_url = ws_url
        self.output_file = output_file
        self.wb_manager = WorkbookManager(output_file)
        self._msg_queue = Queue()
        self._running = False
        self._known_ids = set()

    def start(self):
        """Start the WebSocket listener and REST API fallback poller."""
        self._running = True

        # Start WebSocket thread
        ws_thread = threading.Thread(
            target=self._ws_listener,
            daemon=True,
            name="ws-listener",
        )
        ws_thread.start()

        logger.info(f"[logger] Connecting to {self.ws_url}...")
        logger.info(f"[logger] Excel output: {self.output_file}")
        logger.info(f"[logger] REST API fallback: http://localhost:8000/api/signals")

        last_poll = 0
        rest_url = self.ws_url.rsplit("/", 1)[0].replace("ws:", "http:") + "/api/signals"

        try:
            while self._running:
                try:
                    msg = self._msg_queue.get(timeout=1.0)
                    self._handle_message(msg)
                except Empty:
                    pass

                # REST API fallback every 30 seconds
                now = time.time()
                if now - last_poll >= 30:
                    last_poll = now
                    try:
                        self._rest_poll(rest_url)
                    except Exception as e:
                        logger.debug(f"[rest] Poll error: {e}")

        except KeyboardInterrupt:
            logger.info("[logger] Shutting down...")
            self._running = False

    def _ws_listener(self):
        """Background thread: receive WebSocket messages."""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self._ws_connect())

    async def _ws_connect(self):
        """Async WebSocket connection handler with auto-reconnect."""
        backoff = 1
        while self._running:
            try:
                async with websockets.connect(
                    self.ws_url,
                    ping_interval=20,
                    ping_timeout=20,
                    close_timeout=5,
                    max_size=10 * 1024 * 1024,  # 10MB max message
                ) as ws:
                    logger.info(f"[ws] Connected to {self.ws_url}")
                    backoff = 1  # Reset backoff on successful connect

                    async for raw_msg in ws:
                        try:
                            data = json.loads(raw_msg)
                            self._msg_queue.put(data)
                        except json.JSONDecodeError:
                            logger.warning("[ws] Invalid JSON received")

            except websockets.exceptions.InvalidStatusCode as e:
                logger.error(f"[ws] HTTP error: {e}")
                if not self._running:
                    break
                logger.info(f"[ws] Reconnecting in {backoff}s...")
                time.sleep(backoff)
                backoff = min(backoff * 2, 30)

            except (ConnectionRefusedError, OSError, websockets.exceptions.ConnectionClosed):
                if not self._running:
                    break
                logger.info(f"[ws] Connection lost. Reconnecting in {backoff}s...")
                time.sleep(backoff)
                backoff = min(backoff * 2, 30)

            except Exception as e:
                if not self._running:
                    break
                logger.error(f"[ws] Unexpected error: {e}")
                time.sleep(backoff)
                backoff = min(backoff * 2, 30)

    def _rest_poll(self, url: str):
        """Fallback: poll REST API for signals if WS drops."""
        try:
            import urllib.request
            req = urllib.request.Request(url, headers={"User-Agent": "JudahSignalLogger/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                signals = data.get("signals", [])
                if signals:
                    self._msg_queue.put({"type": "INITIAL", "signals": signals})
                    logger.info(f"[rest] Backfilled {len(signals)} signals from REST API")
        except Exception:
            raise  # caller logs it

    def _handle_message(self, data: dict):
        """Route a message to the appropriate handler."""
        try:
            msg_type = data.get("type", "").upper()
            signals = data.get("signals", [])

            if not signals and msg_type not in ("EXPIRE",):
                return

            if msg_type == "INITIAL":
                logger.info(f"[handler] INITIAL: {len(signals)} signals — creating workbook")
                self.wb_manager.create(signals)
                for s in signals:
                    self._known_ids.add(_v(s, "id"))
                self.wb_manager.rebuild_index()

            elif msg_type == "NEW_SIGNALS":
                new = [s for s in signals if _v(s, "id") not in self._known_ids]
                if new:
                    logger.info(f"[handler] NEW_SIGNALS: {len(new)} new signals")
                    self.wb_manager.add_signals(new)
                    for s in new:
                        self._known_ids.add(_v(s, "id"))

            elif msg_type == "REFRESH":
                logger.debug(f"[handler] REFRESH: {len(signals)} signals updated")
                self.wb_manager.refresh_signals(signals)

            elif msg_type == "EXPIRE":
                logger.info(f"[handler] EXPIRE: {len(signals)} signals expired")
                self.wb_manager.expire_signals(signals)
                for s in signals:
                    self._known_ids.discard(_v(s, "id"))

            elif msg_type == "REVALIDATED":
                logger.debug(f"[handler] REVALIDATED: {len(signals)} signals reset")
                self.wb_manager.refresh_signals(signals)

            else:
                logger.debug(f"[handler] Unknown message type: {msg_type}")

        except Exception as e:
            logger.error(f"[handler] Error processing message: {e}", exc_info=True)


# ─── Entry Point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Real-time Excel logger for Judah Scanner signals"
    )
    parser.add_argument(
        "--ws-url",
        default=os.environ.get("JUDAH_WS_URL", "ws://localhost:8000/ws"),
        help="WebSocket URL (default: ws://localhost:8000/ws)",
    )
    parser.add_argument(
        "--output", "-o",
        default="signal_log.xlsx",
        help="Output Excel file path (default: signal_log.xlsx)",
    )
    args = parser.parse_args()

    # Check dependencies
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    try:
        import websockets  # noqa: F401
    except ImportError:
        print("ERROR: websockets is required. Install with: pip install websockets")
        sys.exit(1)

    logger.info("=" * 60)
    logger.info("  Judah Scanner — Excel Signal Logger")
    logger.info("=" * 60)
    logger.info(f"  WebSocket: {args.ws_url}")
    logger.info(f"  Output:    {args.output}")
    logger.info(f"  Sheets:    Signals | CRT Analysis | SMC Analysis | Trade Levels")
    logger.info("=" * 60)

    app = SignalLogger(args.ws_url, args.output)
    app.start()


if __name__ == "__main__":
    main()
